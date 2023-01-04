import openpyxl
import datetime
import pyexcel
from datetime import datetime as dtime
from datetime import timedelta
from calendar import monthrange
from pathlib import Path
import os
import pandas as pd
from pathManagement import get_current_path,createFolder
import locale

def get_float_from_string(string):
    locale.setlocale(locale.LC_ALL, 'es.UTF-8')
    return locale.atof(string)

def read_cuentas():
    configName="config.xlsx"
    newpath=os.path.join(get_current_path(),configName)
    cuentas=pd.read_excel(newpath,sheet_name="cuentas")
    return cuentas
def read_config():
    bankNamefile=os.path.join(get_current_path(),"config.xlsx")
    sheetName="bancos"
    wb = openpyxl.load_workbook(bankNamefile)
    sheet = wb[sheetName]
    n=sheet.max_row
    m=sheet.max_column
    listBanks=[x[0].value for x in sheet['A2:A'+str(n)]]
    listFields=[x.value for x in sheet['B1:P1'][0]]
    data={}
    for i,bank in enumerate(listBanks):
        data[bank]={}
        for j,field in enumerate(listFields):
            datacell=sheet.cell(row=i+2, column=j+2).value
            data[bank][field]=datacell
    return data
data=read_config()

def get_sheet(bankName,pathFile):
    try:
        wb = openpyxl.load_workbook(pathFile)
    except:
        return None
    #sheet = wb[sheetName]
    sheet=wb.worksheets[0]
    return sheet
def get_dateRow(sheet,i,dateColumn,formatDate):
    if sheet.cell(row=i, column=dateColumn).value==None:
        dateRow=''
    else:
        dateRow=sheet.cell(row=i, column=dateColumn).value

    objectDate=dtime.strptime(str(dateRow),formatDate)
    finalDate=dtime.strftime(objectDate,"%d/%m/%Y")
    return finalDate

def get_documentNr(bankName,sheet,i,documentColumn,codBancaColumn):
    #print("documento :")
    #print(sheet.cell(row=i, column=codBancaColumn).value)
    #print(type(sheet.cell(row=i, column=documentColumn).value))
    if bankName=='Mercantil':
        if sheet.cell(row=i, column=documentColumn).value==None or sheet.cell(row=i, column=documentColumn).value=='':
            if sheet.cell(row=i, column=codBancaColumn).value==None:
                nroDocument=''
                #print("acaaaa")
            else:
                nroDocument=sheet.cell(row=i, column=codBancaColumn).value
                #print("allaaaa")
        else:
            nroDocument=sheet.cell(row=i, column=documentColumn).value
            #print("aquiiiii")
    else:
        if sheet.cell(row=i, column=documentColumn).value==None:
            nroDocument=''
        else:
            nroDocument=sheet.cell(row=i, column=documentColumn).value
    return nroDocument
def get_description(bankName,sheet,i,descriptionColumn,nombreColum,infoComplColumn):
    if bankName=="Mercantil":
        if sheet.cell(row=i, column=descriptionColumn).value==None:
            description=''
        else:
            description=sheet.cell(row=i, column=descriptionColumn).value
        if sheet.cell(row=i, column=nombreColum).value==None:
            name=''
        else:
            name=sheet.cell(row=i, column=nombreColum).value
        description=name+"-"+description+"-"+"Originador ACH"
    elif bankName=="Bisa":
        if sheet.cell(row=i, column=infoComplColumn).value==None:
            description=''
        else:
            description=str(sheet.cell(row=i, column=infoComplColumn).value).replace("Nombre:","")
    else:   
        if sheet.cell(row=i, column=descriptionColumn).value==None:
            description=''
        else:
            description=sheet.cell(row=i, column=descriptionColumn).value
    return description

def get_float_from_string(number):
    indexComma=number.find(",")
    indexPunto=number.find(".")
    if indexComma>=0 and indexPunto>=0:
        #print("hay coma y punto")
        if indexComma>indexPunto:
            #print("hay coma y punto y la coma esta despues del punto")
            number=number.replace(".","").replace(",",".")
            #print(number)
        else:
            #print("hay coma y punto y la coma esta antes del punto")
            number=number.replace(",","")
    elif indexComma>=0 and indexPunto==-1:
        #print("solo hay coma")
        number=number.replace(",",".")
    elif indexPunto>=0 and indexComma==-1:
        pass
        #print("solo hay punto")
    else:
        pass
        #print("no hay coma ni punto")
    return float(number)

def get_number(number):
    #print(type(number))
    if type(number)==float or type(number)==int:
        finalNumber=float(number)
    elif type(number==str):
        finalNumber=get_float_from_string(number)
    pass
    return finalNumber
def get_saldo(sheet,i,descriptionColumn,amountColumn):
    if sheet.cell(row=i, column=amountColumn).value==None:
        saldo=0
    else:
        saldo=get_number(sheet.cell(row=i, column=amountColumn).value)
    return saldo
def get_amount(bankName,sheet,i,amountColumn,creditColumn,debitColumn):
    if bankName=="Mercantil" or bankName=="Fassil":
        if sheet.cell(row=i, column=creditColumn).value==None:
            credit=0
        else:
            credit=get_number(sheet.cell(row=i, column=creditColumn).value)

        if sheet.cell(row=i, column=debitColumn).value==None:
            debit=0
        else:
            debit=get_number(sheet.cell(row=i, column=debitColumn).value)
        amount=credit-debit
    else:
        if sheet.cell(row=i, column=amountColumn).value==None:
            amount=0
        else:
            amount=get_number(sheet.cell(row=i, column=amountColumn).value)
    return amount
def get_typetrx(itfcolumn,amount,sheet,i):
    if sheet.cell(row=i, column=itfcolumn).value==None:
        description=''
    else:
        description=sheet.cell(row=i, column=itfcolumn).value
    
    if description!='':
        if description.find("ITF")!=-1:
            typetrx="ZITF"
        else:
            if amount>0:
                typetrx="Z001"
            else:
                typetrx="Z002"
    else:
        if amount>0:
            typetrx="Z001"
        else:
            typetrx="Z002"
    return typetrx
def maxRowRange(i,sheet,maxRowColumn):
    k=i
    while True:
        if sheet.cell(row=k, column=maxRowColumn).value!=None:
            pass
            k=k+1
        else:
            break
            
    return k-1

def read_bank(fileMeta):
    bankName=fileMeta['bankName']
    pathFile=fileMeta['path']
    sheet = get_sheet(bankName,pathFile)
    if sheet==None:
        return None
    dataUnion=[]

    cuentaColumn=data[bankName]['CuentaCol']
    celdaFecha=data[bankName]['CeldaPeriodo']
    dateColumn=ord(celdaFecha[0])-64 #PARA SALIDA 1
    maxRowColumn=data[bankName]['MaxRowCol']
    documentColumn=data[bankName]['NroDocumCol'] #PARA SALIDA 2 

    creditColumn=data[bankName]['CreditCol']#OPCIONAL PARA SALIDA 3
    debitColumn=data[bankName]['DebitCol'] #OPCIONAL PARA SALIDA 3
    importColumn=data[bankName]['ImporteCol'] # ES PARA SALIDA 3
    saldoColumn=data[bankName]['SaldoCol'] #OPCIONAL PARA SALIDA 3
    nombreColum=data[bankName]['NombreCol']#OPCIONAL PARA SALIDA EN SOLO 1 BANCO 
    descripcionColumn=data[bankName]['DescripcionCol']#OPCIONAL ES PARA SALIDA 4
    codBancaColumn=data[bankName]['CodBancario']#OPCIONAL
    infoCompColumn=data[bankName][' Info. Complementaria']#OPCIONAL
    itfColumn=data[bankName]['CAMPO ITF']#OPCIONAL
    
    dateformat=data[bankName]['FormatoFecha']
    dateExcel=sheet[celdaFecha].value
    #print(f"IMPRIMIENDO FECHA EXCEL VALUE{fileMeta['name']}")
    #print(dateExcel)
    try:
        dateCero = datetime.datetime.strptime(str(dateExcel), dateformat)
    except:
        #write_log("","ERROR EN LA FECHA DEL ARCHIVO",fileMeta['name'])
        raise Exception(f"ERROR EN EL ARCHIVO")
    #print("------------",dateCero,"----------------")
    initialDate=datetime.datetime(dateCero.year,dateCero.month,1)
    finalDate=datetime.datetime(dateCero.year,dateCero.month,monthrange(dateCero.year,dateCero.month)[1])
    i=int(celdaFecha[1:])
    lastRow=maxRowRange(i,sheet,maxRowColumn)
    #print(i,sheet.max_row,descripcionColumn)
    while i<=lastRow:
        if sheet.cell(row=i, column=maxRowColumn).value!=None:
            dateRow=get_dateRow(sheet,i,dateColumn,dateformat)
            nroDocument=get_documentNr(bankName,sheet,i,documentColumn,codBancaColumn)
            saldo=get_saldo(sheet,i,saldoColumn,saldoColumn)
            description=get_description(bankName,sheet,i,descripcionColumn,nombreColum,infoCompColumn) #DINAMICO 
            amount=get_amount(bankName,sheet,i,importColumn,creditColumn,debitColumn) #DINAMICO
            typetrx=get_typetrx(itfColumn,amount,sheet,i)
            unionRow={
                "date":dateRow,
                "documentNr":nroDocument,
                "description":description,
                "type":typetrx,
                "amount":amount,
                "saldo":saldo,
            }
            dataUnion.append(unionRow)
        i=i+1
    initialBalance=dataUnion[0]['saldo']-dataUnion[0]['amount']
    finalBalance=dataUnion[len(dataUnion)-1]['saldo']
    #print(initialBalance,finalBalance)
    #print(data[bankName])
    return {'data':dataUnion,'initialBalance':initialBalance,
    'finalBalance':finalBalance,'initialDate':initialDate,
    'finalDate':finalDate,'account':fileMeta['name'][:4],'namComercial':data[bankName]['NombreComercial']}
def get_template_base():
    filename="plantillaSap.xlsx"
    extractFilePath=os.path.join(get_current_path(), "plantillasSap",filename)
    wb = openpyxl.load_workbook(extractFilePath)
    return wb
def make_templates(infobank):
    #print(infobank)
    wb=get_template_base()
    sheet = wb["UNION"]
    initialDate=infobank['initialDate'].strftime("%d/%m/%Y")
    finalDate=infobank['finalDate'].strftime("%d/%m/%Y")
    initialBalance=infobank['initialBalance']
    finalBalance=infobank['finalBalance']
    accountNumber=infobank['account']
    newPeriod=f"ESTADO DE CUENTA DEL {initialDate} AL {finalDate}"
    dataSap=infobank['data']
    binAccountouput=infobank['account'][:4]
    sheet["A1"]=newPeriod
    sheet["C4"]=accountNumber
    sheet["A9"]=initialBalance
    sheet["B9"]=finalBalance
    sheet["A8"]=initialDate
    sheet["B8"]=finalDate
    sheet["B3"]=infobank['namComercial']
    rowInit=17
    
    for i in range(len(dataSap)):
        sheet.cell(row=rowInit,column=1).value=dataSap[i]['date']
        sheet.cell(row=rowInit,column=2).value=dataSap[i]['documentNr']
        sheet.cell(row=rowInit,column=3).value=dataSap[i]['description']
        sheet.cell(row=rowInit,column=4).value=dataSap[i]['type']
        sheet.cell(row=rowInit,column=5).value=dataSap[i]['amount']
        rowInit=rowInit+1
    
    binAccount=infobank['account'][-4:]
    dateFname=datetime.datetime.now().date()
    daybefore=dateFname-datetime.timedelta(days=1)
    dateFname=daybefore.strftime("%d%m%Y")
    fileTemplateName=f"{binAccount}-{dateFname}.xlsx"
    fileTemplatePath=os.path.join(get_current_path(), "plantillasSap",dateFname,fileTemplateName)
    fileTemplatefolder=os.path.join(get_current_path(), "plantillasSap",dateFname)
    wb.save(fileTemplatePath)
    #print(f"SE CREO EL ARCHIVO {fileTemplateName}")
    #print(fileTemplatefolder)
    #write_log(" ","OK",fileTemplatefolder)

def convert_xls(pathFolder):    
    filesInfolder=os.listdir(pathFolder)
    e=""
    for file in filesInfolder:
        if file.endswith(".xls"):
            try:
                print(file)
                xls=pathFolder+"\\"+file
                xlsx=pathFolder+"\\"+file.replace(".xls",".xlsx")
                #print(xls)
                #print(xlsx)
                pyexcel.save_book_as(file_name=xls, dest_file_name=xlsx)
            except Exception as e:
                print(e)
                os.remove(xlsx)
                #write_log(f"")
            #os.remove(xls)
    return e
def get_extrac_files():
    # get the current date
    tableAcounts=read_cuentas().values.tolist()
    currentFolder=datetime.datetime.now().date()
    dayBefore=currentFolder-datetime.timedelta(days=1)
    dayBefore=dayBefore.strftime("%d%m%Y")
    newpath=os.path.join(get_current_path(), "extractosBancarios",dayBefore)
    e=convert_xls(newpath)
    files = [f for f in os.listdir(newpath) if f.endswith('.xlsx')]
    #files = os.listdir(newpath)
    paths=[]
    for f in files:
        for tb in tableAcounts:
            #print(str(tb[3])[-4:],"-----",f[:4])
            if f[:4]==str(tb[3])[-4:]:
                society=tb[4]
                codeBank=tb[5]
                currency=tb[6]
                bankName=tb[2]
                break
        filemeta={
            "name":f,
            "path":os.path.join(newpath,f),
            "society":society,
            "codeBank":codeBank,
            "currency":currency,
            "bankName":bankName,
            "characterNovalid":e,
        }
        #print(f[:4])
        paths.append(filemeta)
    return paths
def write_log(s,log,rut):
    txtfolder=os.path.dirname(rut)
    pathLog=os.path.join(txtfolder, "logs.txt")
    line=s+str(log)
    if s=="" and log!="\n":
        line=line+" "+datetime.datetime.today().strftime("%d/%m/%Y %H:%M:%S")
    print(log)
    with open(pathLog, 'a') as file:
        file.write(line)
def process_xlsxFiles():
    currentFolder=datetime.datetime.now().date()
    dayBefore=currentFolder-datetime.timedelta(days=1)
    dayBefore=dayBefore.strftime("%d%m%Y")
    folderExist=createFolder(os.path.join(get_current_path(), "extractosBancarios",dayBefore),force=False,delete=False)
    if not(folderExist):
        print("EL FOLDER DIARIO NO EXISTE O TIENE FORMATO INCORRECTO",dayBefore)
        return False
    createFolder(os.path.join(get_current_path(), "plantillasSap",dayBefore),force=True,delete=True)
    pathFiles=get_extrac_files()
    print(len(pathFiles))
    r=0
    for fileMeta in pathFiles:
        try:
            log=""+fileMeta['name']
            write_log("",log,fileMeta['path'])
            infobank=read_bank(fileMeta)
            if infobank==None:
                log="ARCHIVO NO PROCESADO ERROR AL ABRIR EL ARCHIVO"
                write_log(" ",log,fileMeta['path'])
            else:
                make_templates(infobank)
                r=r+1
                write_log(" ","OK",fileMeta['path'])
        except Exception as e:
            write_log(" ",e,fileMeta['path'])
        write_log("","\n",fileMeta['path'])
    try:
        ratio="{:.2f}".format(r/len(pathFiles)*100)
    except:
        ratio=0
    results=f"\n***************  PROCESO TERMINADO AL {ratio}%  ***************"
    try:
        write_log("",results,pathFiles[0]['path'])
    except:
        pass
    return True
if __name__ == "__main__":
    process_xlsxFiles()